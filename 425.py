
from dwave.system import LeapHybridSampler
import numpy as np
import time
from dwave.system import DWaveSampler, EmbeddingComposite
from dimod import BinaryQuadraticModel
import openpyxl as pyxl
import matplotlib
try:
    import matplotlib.pyplot as plt
except ImportError:
    matplotlib.use("agg")
    import matplotlib.pyplot as plt
from matplotlib import animation

class Parameter(object):

    def __init__(self, filename):
        # System
        # 1. System Data
        tool = Excel_tool()
        data = tool.read_excel(filename)
        # 1) line
        self.Line = data[0]
        self.N_line = len(self.Line)  # number of line
        # 2) bus
        self.Bus = data[1]
        self.N_bus = len(self.Bus)  # number of bus
        # 3) substation
        self.Sub = data[2]
        self.N_sub = len(self.Sub)  # number of substation
        # # 4) generator
        # self.Gen = data[3]
        # self.N_gen = len(self.Gen)  # number of renewables
        # self.Factor = data[1, 6]  # power factor (rad)

        # 4) daily curve
        self.LoadDay = data[3]
        self.N_time = len(self.LoadDay)  # number of hours

        # 7) MGO_DG power generation
        self.MGO_DG = data[4]
        self.N_MGO = len(self.MGO_DG)  # number of
        self.N_cDG = len(self.MGO_DG)

        # 7) MGO_ESS power generation
        self.MGO_ESS = data[5]
        self.N_ESS = len(self.MGO_ESS)  # number of

        # 7) MGO_DR power generation
        self.MGO_DR = data[6]
        self.N_DR = len(self.MGO_DR)  # number of

        # 6) global variable index
        self.var_index()

        # Base value
        self.Base_V = 12.66  # voltage: 12.66 kV
        self.Base_S = 10.00  # power:   10.00 MVA

        # Cost
        self.Cost_sub = 83  # cost of power purchasing
        self.Cost_pen = data[1]  # cost of load shedding

        self.Cost_los = 25  # cost of power loss
        # Other
        self.Big_M = 100  # a sufficient large number
        self.V_min = (0.9 * self.Base_V)
        self.V_max = (1.1 * self.Base_V)
        # Bus-Line Information
        self.Line_head = [[] for i in range(self.N_bus)]
        self.Line_tail = [[] for i in range(self.N_bus)]
        for i in range(self.N_line):
            head = self.Line[i][1]
            tail = self.Line[i][2]
            self.Line_head[int(round(head))].append(i)
            self.Line_tail[int(round(tail))].append(i)

        # Line－bus Information
        self.Bus_head = [[] for i in range(self.N_line)]
        self.Bus_tail = [[] for i in range(self.N_line)]
        for i in range(self.N_line):
            bhead = self.Line[i][1]
            btail = self.Line[i][2]
            self.Bus_head[i].append(int(round(bhead)))
            self.Bus_tail[i].append(int(round(btail)))



        self.damage = data[7]
        self.periods = 12
        self.tt = np.array([[0, 1, 2, 2, 2, 3, 3, 4, 4],
                            [1, 0, 1, 2, 1, 3, 2, 3, 3],
                            [2, 1, 0, 3, 2, 4, 1, 4, 4],
                            [2, 2, 3, 0, 3, 1, 4, 2, 5],
                            [2, 1, 2, 3, 0, 2, 1, 2, 2],
                            [3, 3, 4, 1, 2, 0, 3, 1, 4],
                            [3, 2, 1, 4, 1, 3, 0, 3, 3],
                            [4, 3, 4, 2, 2, 1, 3, 0, 4],
                            [4, 3, 4, 5, 2, 4, 3, 4, 0]])
        self.DN1 = np.array([7, 1, 3, 4, 5, 7])  # 和ID不同，后面是节点编号， ID为线路编号
        self.rt = np.array([[0, 1, 1, 1, 1, 0],
                            [0, 1, 1, 1, 1, 0]])
    # This function creates global index
    def var_index(self):
        # Fictitious power flow
        # 1. Name
        # global N_F_line, N_F_sub, N_F_load, N_F_gen, N_F_MGO, N_F_var
        # # 2. Initialization
        # N_F_line = 0  # flow of line
        # N_F_load = N_F_line + self.N_line  # flow of load demand
        # N_F_sub = N_F_load + self.N_bus  # flow of substation
        # N_F_gen = N_F_sub + self.N_sub  # flow of DG
        # N_F_MGO = N_F_gen + self.N_gen  # flow of MGO
        # N_F_var = N_F_MGO + self.N_MGO  # Number of all variables

        # Real power flow
        # 1. Name
        global N_V_bus, N_I_line, N_P_line, N_Q_line, N_P_sub, N_Q_sub
        global N_S_gen, N_C_gen, N_P_cut, N_Q_cut, N_P_MGO, N_Q_MGO
        global N_P_cDG, N_Q_cDG, N_P_ESSd, N_P_ESSc, N_E_ESS, N_P_DR1, N_P_DR2, N_P_DR, N_N_var
        # 2. Initialization
        N_V_bus = 0  # square of voltage amplitude
        # N_I_line = N_V_bus + self.N_bus  # square of voltage phase angle
        # N_P_line = N_I_line + self.N_line  # power flow (active)
        N_P_line = N_V_bus + self.N_bus  # power flow (active)
        N_Q_line = N_P_line + self.N_line  # power flow (reactive)
        N_P_sub = N_Q_line + self.N_line  # power injection at substation
        N_Q_sub = N_P_sub + self.N_sub  # power injection at substation
        N_P_cut = N_Q_sub + self.N_sub  # Load shedding (active)
        N_Q_cut = N_P_cut + self.N_bus  # Load shedding (reactive)
        N_P_MGO = N_Q_cut + self.N_bus  # MGO (active)
        N_Q_MGO = N_P_MGO + self.N_MGO  # MGO (reactive)
        N_P_cDG = N_Q_MGO + self.N_MGO  # cDG in MGO (active)
        N_Q_cDG = N_P_cDG + self.N_cDG  # cDG in MGO (active)
        N_P_ESSd = N_Q_cDG + self.N_cDG  # ESSd in MGo (active)
        N_P_ESSc = N_P_ESSd + self.N_ESS  # ESSc in MGo (active)
        N_E_ESS = N_P_ESSc + self.N_ESS # ESSe in MGo (active)
        # N_P_DR1 = N_E_ESS + self.N_ESS # DR1 in DR(active)
        # N_P_DR2 = N_P_DR1 + self.N_DR  # DR2 in DR(active)
        N_P_DR = N_E_ESS + self.N_DR   # DR in DR(active)
        N_N_var = N_P_DR + self.N_DR  # Number of all variables

# This class creates the execution tool for Excel files
class Excel_tool(object):

    # Initialization
    def __init__(self):
        pass

    # inputs data from Excel file
    def read_excel(self, filename):
        data = []
        book = pyxl.load_workbook(filename)
        # Data preprocessing
        for i, name in enumerate(book.sheetnames):  # sheet number
            if i < len(book.sheetnames):
                sheet = book[name]
                n_row = sheet.max_row  # number of rows
                n_col = sheet.max_column  # number of columns
                data.append(self.tool_filter(sheet, n_row, n_col))
        return data

    # saving data to excel file
    def save_excel(self, filename, sheetname, title, data):
        book = pyxl.load_workbook(filename)
        name = book.sheetnames[-1]
        book.remove(book[name])
        sheet = book.create_sheet(sheetname)
        sheet.append(title)
        for i in range(len(data)):
            sheet.append(data[i, :].tolist())  # write data
        book.save(filename=filename)

    # filter data from the original numpy array
    def tool_filter(self, sheet, n_row, n_col):
        k = 0
        data = []
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if it is a number
                data.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position
                    val = sheet[pos].value
                    if sheet[pos].data_type == 'n':
                        data[k].append(val)
                k = k + 1
        return np.array(data)

def build_bqm(Para, added_cuts):
    """Build bqm that models our problem.
    Args:


    Returns:
        bqm (BinaryQuadraticModel): QUBO model for the input scenario
        x (list of strings): List of variable names in BQM
    """

    print("\nBuilding binary quadratic model...")

    # initialize the parameters
    fault_line = np.zeros((10, 2), dtype=int)
    for l in list([0, 2, 3, 4]):
        fault_line[l, 0] = Para.Line[l, 1]
        fault_line[l, 1] = Para.Line[l, 2]
    Root_node_potential_info = list(fault_line.reshape(-1)) + list([0, 1, 5, 6])  # 损坏线路+ MGO+ 配网主节点
    Root_node_potential = list(set([i for i in Root_node_potential_info]))


    # Build variables
    x1_fix = [[['x' + str(c) + '_' + str(d) + '_' + str(s) for s in range(6)] for d in range(6)] for c in range(2)]
    y1_fix = [[ 'y' + str(c) + '_' + str(d) for d in range(6)] for c in range(2)]
    T = [[ 'T' + str(d) + '_' + str(t) for t in range(Para.periods)] for d in range(6)]
    y_line = [[ 'yline' + str(l) + '_' + str(t) for t in range(Para.periods)] for l in range(Para.N_line)]
    gama = [[ 'gama' + str(i) + '_' + str(t) for t in range(Para.periods)] for i in range(len(Root_node_potential))]
    z_fix = [[ 'z_fix' + str(d) + '_' + str(t) for t in range(Para.periods)] for d in range(6)]
    Q = [[ 'Q' + str(d) + '_' + str(i) for i in range(10)] for d in range(6)] # integer 离散为10个binary变量
    f = [[['f' + str(l) + '_' + str(t) + '_' + str(i) for i in range(10)] for t in range(Para.periods)] for l in range(Para.N_line)  ]# integer 离散为10个binary变量
    AT = [[['AT' + str(c) + '_' + str(d) + '_' + str(i) for i in range(10)] for d in range(6)] for c in range(2)  ]# integer 离散为10个binary变量

    hw = ['hw' + str(i) for i in range(20)]

    # 定义inequality_constraint



    # time
    time = list(range(0, 24))
    root = list(range(len(Root_node_potential)))

    # Initialize BQM
    bqm = BinaryQuadraticModel('BINARY')

    # Objective
    alpha_pair = 2
    for d in range(1, 5):  #
        for t in range(Para.periods):
            bqm.add_variable(T[d][t], (t + 1) * alpha_pair)

    for i in range(20):
        bqm.add_variable(hw[i], 2** (i - 10))

    # Constraint: topology reconfiguration constraints
    for t in range(Para.periods):
        for m in range(Para.N_bus):
            if m not in Root_node_potential:
                c1 = [(f[l][t][i], i) for i in range(10) for l in range(Para.N_line) if Para.Line[l, 2] == m]
                c2 = [(f[l][t][i], -i) for i in range(10) for l in range(Para.N_line) if Para.Line[l, 1] == m]
                bqm.add_linear_equality_constraint(c1 + c2, 1, -1)

    for t in range(Para.periods):
        for m in range(len(Root_node_potential)):
            c3 = [(f[l][t][i], i) for i in range(10) for l in range(Para.N_line) if
                  Para.Line[l, 2] == Root_node_potential[m]]
            c4 = [(f[l][t][i], -i) for i in range(10) for l in range(Para.N_line) if
                  Para.Line[l, 1] == Root_node_potential[m]]
            c5 = [(gama[m][t], (-Para.Big_M))]
            bqm.add_linear_inequality_constraint(c3 + c4 + c5,
                                                 lb=np.iinfo(np.int64).min,
                                                 ub=0,
                                                 lagrange_multiplier=1,
                                                 label='con2/1_' + str(m) + str(t),
                                                 constant=-1)

    for t in range(Para.periods):
        for m in range(len(Root_node_potential)):
            c6 = [(f[l][t][i], -i) for i in range(10) for l in range(Para.N_line) if
                  Para.Line[l, 2] == Root_node_potential[m]]
            c7 = [(f[l][t][i], i) for i in range(10) for l in range(Para.N_line) if
                  Para.Line[l, 1] == Root_node_potential[m]]
            c8 = [(gama[m][t], (-Para.Big_M))]
            bqm.add_linear_inequality_constraint(c6 + c7 + c8,
                                                 constant=1,
                                                 lagrange_multiplier=1,
                                                 label='con2/2_' + str(m) + str(t))

    for t in range(Para.periods):
        for l in range(Para.N_line):
            c9 = [(f[l][t][i], i) for i in range(10)]
            c99 = [(y_line[l][t], (-Para.Big_M))]
            bqm.add_linear_inequality_constraint(c9 + c99,
                                                 constant=0,
                                                 lagrange_multiplier=1,
                                                 label='con3')

    for t in range(Para.periods):
        for l in range(Para.N_line):
            c10 = [(f[l][t][i], -i) for i in range(10)]
            c100 = [(y_line[l][t], (-Para.Big_M))]
            bqm.add_linear_inequality_constraint(c10 + c100,
                                                 constant=0,
                                                 lagrange_multiplier=1,
                                                 label='con4')

    for t in range(Para.periods):
        c11 = [(y_line[l][t], 1) for l in range(Para.N_line)]
        c12 = [(gama[i][t], 1) for i in range(len(Root_node_potential))]
        bqm.add_linear_equality_constraint(c11 + c12, 1, -Para.N_bus)

    # Constraint: repair dispatch constraints
    damageline_ID = np.array([7, 0, 2, 3, 4])  # damage line and depot
    ndn = len(damageline_ID)

    for c in range(2):
        for d in range(0, ndn + 1):
            if (d == 0):
                c13 = [(x1_fix[c][d][i], 1) for i in range(0, ndn)]
                c14 = [(x1_fix[c][d][d], -1)]
                c15 = [(y1_fix[c][d], -1)]
                bqm.add_linear_equality_constraint(c13 + c14 + c15, 1, 0)
            elif (d == 5):
                c16 = [(y1_fix[c][d], 1)]
                bqm.add_linear_equality_constraint(c16, 1, -1)
            else:
                c17 = [(x1_fix[c][d][i], 1) for i in range(1, ndn + 1)]
                c18 = [(x1_fix[c][d][d], -1)]
                c19 = [(y1_fix[c][d], -1)]
                bqm.add_linear_equality_constraint(c17 + c18 + c19, 1, 0)

    for c in range(2):
        for d in range(1, ndn):
            c20 = [(x1_fix[c][d][i], 1) for i in range(1, ndn + 1)]
            c21 = [(x1_fix[c][i][d], -1) for i in range(0, ndn)]
            bqm.add_linear_equality_constraint(c20 + c21, 1, 0)

    for d in range(1, ndn):
        bqm.add_linear_equality_constraint([(y1_fix[0][d], 1), (y1_fix[1][d], 1)], 1, -1)

    for c in range(2):
        c22 = [(x1_fix[c][0][i], 1) for i in range(1, ndn)]
        bqm.add_linear_equality_constraint(c22, 1, -1)

    for c in range(2):
        c23 = [(x1_fix[c][i][5], 1) for i in range(1, ndn)]
        bqm.add_linear_equality_constraint(c23, 1, -1)

    ## Q

    c24 = [(Q[0][i], i + 1) for i in range(10)]
    bqm.add_linear_equality_constraint(c24, 1, -1)

    for i in range(ndn):
        c25 = [(Q[i][j], -(j + 1)) for j in range(10)]
        bqm.add_linear_inequality_constraint(c25,
                                             constant=0,
                                             lagrange_multiplier=1,
                                             label='con5')

    for c in range(2):
        for d in range(0, ndn):
            for s in range(1, ndn + 1):
                if not (s == d):
                    c26 = [(Q[d][i], i + 1) for i in range(10)]
                    c27 = [(Q[s][i], -(i + 1)) for i in range(10)]
                    c28 = [(x1_fix[c][d][s], ndn - 1)]
                    bqm.add_linear_inequality_constraint(c26 + c27 + c28,
                                                         constant=-ndn + 2,
                                                         lagrange_multiplier=1,
                                                         label='con6')

    for c in range(2):
        for d in range(0, ndn):
            for s in range(1, ndn + 1):
                dd = Para.DN1[d]
                ss = Para.DN1[s]
                if not (s == d):
                    c29 = [(AT[c][d][i], i) for i in range(10)]
                    c30 = [(AT[c][s][i], -i) for i in range(10)]
                    c31 = [(x1_fix[c][d][s], 100)]
                    bqm.add_linear_inequality_constraint(c29 + c30 + c31,
                                                         constant=Para.rt[c, d] + Para.tt[dd, ss] - 100,
                                                         lagrange_multiplier=1,
                                                         label='con7')

    for c in range(2):
        for d in range(1, ndn):
            c32 = [(AT[c][d][i], -i) for i in range(10)]
            c33 = [(AT[c][d][i], i) for i in range(10)]
            c34 = [(y1_fix[c][d], -100)]
            bqm.add_linear_inequality_constraint(c32,
                                                 constant=0,
                                                 lagrange_multiplier=1,
                                                 label='con8')
            bqm.add_linear_inequality_constraint(c33 + c34,
                                                 constant=0,
                                                 lagrange_multiplier=1,
                                                 label='con9')

    for d in range(1, ndn):
        c35 = [(T[d][t], 1) for t in range(Para.periods)]
        bqm.add_linear_equality_constraint(c35, 1, -1)

    for d in range(1, ndn):
        c36 = [(AT[c][d][i], i) for c in range(2) for i in range(10)]
        c37 = [(y1_fix[c][d], Para.rt[c][d]) for c in range(2)]
        c38 = [(T[d][t], -(t + 1)) for t in range(Para.periods)]
        bqm.add_linear_inequality_constraint(c36 + c37 + c38,
                                             constant=0,
                                             lagrange_multiplier=1,
                                             label='con10')

        c39 = [(AT[c][d][i], -i) for c in range(2) for i in range(10)]
        c40 = [(y1_fix[c][d], -Para.rt[c][d]) for c in range(2)]
        c41 = [(T[d][t], (t + 1)) for t in range(Para.periods)]
        bqm.add_linear_inequality_constraint(c39 + c40 + c41,
                                             constant=1 - 0.0001,
                                             lagrange_multiplier=1,
                                             label='con11')

    for t in range(Para.periods):
        for d in range(0, ndn + 1):
            c42 = [(T[d][i], -1) for i in range(t)]
            c43 = [(z_fix[d][t], 1)]
            bqm.add_linear_equality_constraint(c42 + c43, 1, 0)

    for d in damageline_ID:  # ID number of damaged lines
        n = int(np.where(damageline_ID == d)[0])
        if not (n == 0):
            for t in range(Para.periods):
                bqm.add_linear_inequality_constraint([(y_line[d][t], 1), (z_fix[n][t], -1)],
                                                     constant=0,
                                                     lagrange_multiplier=1,
                                                     label='con12')
    # start处理上次迭代传送回来的added_cuts：—————————————————————————————————————————————————————————————————————————
    for item in added_cuts:  # [0, sub_ray1, sub_ray2, sub_ray3, sub_ray4, sub_ray5, sub_ray6, sub_ray7, sub_ray8, sub_ray9, sub_ray10, sub_ray11, sub_ray12, sub_ray13, sub_ray14, sub_ray15, sub_ray16, sub_ray17, sub_ray18, sub_ray19, sub_ray20, sub_ray191, sub_ray201, sub_ray21, sub_ray22, sub_ray23, sub_ray24, sub_ray25, sub_ray26, sub_ray27, sub_ray28, sub_rayl1, sub_rayl2, sub_rayl3, sub_rayl4, sub_rayl5, sub_rayl6, sub_rayl7]
        #  对偶形式的目标函数 ————————————————————————————————————————————————————————————————————————————————————————

        cut_term = 0
        #  目标函数中f_T*lambda部分——————————————————————————————————————————————————————————————————————————————————
        for t in range(Para.periods):
            for n in range(Para.N_bus):
                cut_term = cut_term + item[31][n, t] * Para.Bus[n, 1] * Para.LoadDay[t, 1]

        for t in range(Para.periods):
            for n in range(Para.N_bus):
                cut_term = cut_term + item[32][n, t] * Para.Bus[n, 2] * Para.LoadDay[t, 1]

        for i in range(Para.N_MGO):
            cut_term = cut_term + item[34][i, 0] * 3

        for n in range(Para.N_DR):
            cut_term = cut_term + item[35][n, 0] * Para.MGO_DR[n, 4]

        for t in range(Para.periods):
            for n in range(Para.N_MGO):
                cut_term = cut_term + item[36][n, t] * (0.5 * Para.LoadDay[t, 2] + 0.5 * Para.LoadDay[t, 3])

        # 目标函数中 （g-Gx）_T * alpha 部分———————————————————————————————————————————————————————————————————————
        for t in range(Para.periods):  # alpha 2
            for n in range(Para.N_bus):
                cut_term = cut_term + item[2][n, t] * Para.Bus[n, 1] * Para.LoadDay[t, 1]

        for t in range(Para.periods):  # alpha 9
            for n in range(Para.N_bus):
                cut_term = cut_term + item[9][n, t] * Para.V_min

        for t in range(Para.periods):  # alpha 10
            for n in range(Para.N_bus):
                cut_term = cut_term + item[10][n, t] * Para.V_max

        for t in range(Para.periods):  # alpha 11
            cut_term = cut_term + item[11][t] * Para.Sub[0, 5]
        for t in range(Para.periods):  # alpha 12
            cut_term = cut_term + item[12][t] * Para.Sub[0, 4]
        for t in range(Para.periods):  # alpha 13
            cut_term = cut_term + item[13][t] * Para.Sub[0, 3]
        for t in range(Para.periods):  # alpha 14
            cut_term = cut_term + item[14][t] * Para.Sub[0, 2]

        PMGOmax = 0.5
        PMGOmin = -0.5
        QMGOmax = 0.5
        QMGOmin = -0.5
        for t in range(Para.periods):  # alpha 15
            for n in range(Para.N_MGO):
                cut_term = cut_term + item[15][n, t] * PMGOmin

        for t in range(Para.periods):  # alpha 16
            for n in range(Para.N_MGO):
                cut_term = cut_term + item[16][n, t] * PMGOmax

        for t in range(Para.periods):  # alpha 17
            for n in range(Para.N_MGO):
                cut_term = cut_term + item[17][n, t] * QMGOmin

        for t in range(Para.periods):  # alpha 18
            for n in range(Para.N_MGO):
                cut_term = cut_term + item[18][n, t] * QMGOmax

        for t in range(Para.periods):  # alpha 19
            for n in range(Para.N_cDG):
                cut_term = cut_term + item[19][n, t] * Para.MGO_DG[n, 3]

        for t in range(Para.periods):  # alpha 20
            for n in range(Para.N_cDG):
                cut_term = cut_term + item[20][n, t] * Para.MGO_DG[n, 2]

        for t in range(Para.periods):  # alpha 191
            for n in range(Para.N_cDG):
                cut_term = cut_term + item[21][n, t] * Para.MGO_DG[n, 5]

        for t in range(Para.periods):  # alpha 201
            for n in range(Para.N_cDG):
                cut_term = cut_term + item[22][n, t] * Para.MGO_DG[n, 4]

        for t in range(Para.periods):  # alpha 21
            for n in range(Para.N_ESS):
                cut_term = cut_term + item[23][n, t] * Para.MGO_ESS[n, 3]

        for t in range(Para.periods):  # alpha 22
            for n in range(Para.N_ESS):
                cut_term = cut_term + item[24][n, t] * Para.MGO_ESS[n, 2]

        for t in range(Para.periods):  # alpha 23
            for n in range(Para.N_ESS):
                cut_term = cut_term + item[25][n, t] * Para.MGO_ESS[n, 3]

        for t in range(Para.periods):  # alpha 24
            for n in range(Para.N_ESS):
                cut_term = cut_term + item[26][n, t] * Para.MGO_ESS[n, 2]

        for t in range(Para.periods - 1):  # alpha 25
            for n in range(Para.N_ESS):
                cut_term = cut_term + item[27][n, t] * Para.MGO_ESS[n, 5] * Para.MGO_ESS[n, 6]

        for t in range(Para.periods - 1):  # alpha 26
            for n in range(Para.N_ESS):
                cut_term = cut_term + item[28][n, t] * Para.MGO_ESS[n, 4] * Para.MGO_ESS[n, 6]

        for t in range(Para.periods):  # alpha 27
            for n in range(Para.N_DR):
                cut_term = cut_term + item[29][n, t] * Para.MGO_DR[n, 3]

        for t in range(Para.periods):  # alpha 28
            for n in range(Para.N_DR):
                cut_term = cut_term + item[30][n, t] * Para.MGO_DR[n, 2]

        # end处理上次迭代传送回来的added_cuts：—————————————————————————————————————————————————————————————————————————

        c44 = [(y_line[n][t], item[3][n, t] * Para.Big_M) for t in range(Para.periods) for n in range(Para.N_line)]
        c45 = [(y_line[n][t], -item[4][n, t] * Para.Big_M) for t in range(Para.periods) for n in range(Para.N_line)]
        c46 = [(y_line[n][t], -item[5][n, t] * Para.Line[n, 3]) for t in range(Para.periods) for n in
               range(Para.N_line)]
        c47 = [(y_line[n][t], item[6][n, t] * Para.Line[n, 3]) for t in range(Para.periods) for n in range(Para.N_line)]
        c48 = [(y_line[n][t], -item[7][n, t] * Para.Line[n, 3]) for t in range(Para.periods) for n in
               range(Para.N_line)]
        c49 = [(y_line[n][t], item[8][n, t] * Para.Line[n, 3]) for t in range(Para.periods) for n in range(Para.N_line)]
        c50 = [(hw[i], -2 ** (i - 10)) for i in range(20)]

        for t in range(Para.periods):
            for n in range(Para.N_line):
                cut_term = cut_term - item[3][n, t] * Para.Big_M + item[4][n, t] * Para.Big_M

        bqm.add_linear_inequality_constraint(c44 + c45 + c46 + c47 + c48 + c49 + c50,
                                             constant=cut_term,
                                             lagrange_multiplier=1,
                                             label='con13')

    return bqm, y_line, gama, x1_fix, y1_fix, T, AT, Q, f, hw


if __name__ == "__main__":
    added_cuts = [[1, np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [-10., -10., -10., -10., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [-10., -10., -10., -10., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., -10., 0., -10., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [-20., -20., -20., -20., -20., -20., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [-20., -20., -20., -20., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.]]), np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., -0., -0., -0., -0., -0., 0., 0.],
                             [0., 0., 0., 0., 0., -0., 0., 0., -0., 0., 0., 0.],
                             [0., 0., 0., 0., -0., -0., 0., 0., -0., 0., -0., -0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., -0., -0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., -0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [-0., -0., 0., -0., -0., -0., -0., 0., 0., 0., 0., 0.],
                             [0., -0., 0., -0., -0., 0., 0., -0., -0., -0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., -0., -0., -0., 0., 0.],
                             [0., -0., 0., -0., -0., -0., 0., -0., 0., 0., -0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [10., 0., 10., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 100., 100., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 100., 100., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.]]), np.array([[-100., -100., -100., -100., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., -0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., -100., -100., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.],
                                               [-10., 0., -10., 0., 0., 0., 0., 0., 0.,
                                                0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., -0., -0., 0., 0., -0., 0., 0., -0., -0., -0., -0.],
                             [-0., -0., -0., -0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., -0., -0., -0., -0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., -0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., -0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [-0., 0., 0., 0., 0., -0., 0., 0., 0., 0., 0., 0.]]), np.array([[0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.],
                                                                                             [0.]]), np.array([[0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.],
                                                                                                               [0.]]),
                   np.array([[0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.],
                             [0.]]), np.array([[0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.],
                                               [0.]]), np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                                                 [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                                                 [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[-100., -100., -100., -100., -0., 0., 0., -0., -0.,
                              -0., 0., 0.],
                             [-100., -100., -100., -100., -0., 0., 0., 0., 0.,
                              0., 0., -0.],
                             [-100., -100., -100., -100., -0., 0., 0., 0., 0.,
                              0., 0., 0.]]), np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                                       [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                                       [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., -0., -0., 0., 0., -0., -0., -0., -0.],
                             [-0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., -0., 0., -0., -0., -0., 0.],
                             [-0., -0., -0., -0., -0., 0., 0., 0., 0., 0., -0., 0.],
                             [-0., 0., -0., -0., 0., 0., 0., 0., -0., -0., -0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[-0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0.],
                             [-0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0.],
                             [-0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0., -0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., -0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., -0., 0., -0., -0., -0., -0.],
                             [0., 0., -0., -0., 0., 0., 0., 0., 0., 0., -0., -0.],
                             [0., 0., 0., -0., 0., 0., 0., 0., -0., -0., -0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [90., 100., 90., 100., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 100., 100., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 0., 0., 0., 0., 0., 0., 0.,
                              0.],
                             [100., 100., 100., 100., 100., 100., 0., 0., 0., 0., 0.,
                              0.]]), np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                                               [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]]), np.array([[0.],
                                                                                       [0.],
                                                                                       [0.]]),
                   np.array([[0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.],
                             [0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0., 0.]])]]
    Para = Parameter("Data-9bus.xlsx")
    tool = Excel_tool()

    bqm, y_line, gama, x1_fix, y1_fix, T, AT, Q, f, hw = build_bqm(Para, added_cuts)
    print("\nRunning QPU solver...")
    # 使用 Direct QPU Solver 求解器_______________________________
    sampler = EmbeddingComposite(DWaveSampler(solver='Advantage_system4.1'))
    qubo = bqm.to_qubo()
    sampleset = sampler.sample_qubo(qubo[0], num_reads=1, label='direct QPU' + str(0))
    sample = sampleset.first.sample
    print(0)